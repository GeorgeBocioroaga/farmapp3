from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Body
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import date
from pydantic import ValidationError
from db import get_db
from models import (
    Inventory,
    InventoryMovement,
    ChemMixRule,
    Doc,
    ActiveSubstance,
    ChemProduct,
    ProductActive,
    InventoryLocation,
    StockLot,
    InventoryTxn,
    Application,
    ApplicationItem,
)
from schemas import (
    InventoryCreate,
    InventoryUpdate,
    InventoryMovementCreate,
    MixCheckRequest,
    ChemProductUpsert,
    InventoryLotCreate,
    InventoryTxnCreate,
    ActiveSubstanceCreate,
)
from security import get_current_user
from services import chem_parse, storage, chem_units
import requests
import time
import os
from openpyxl import Workbook
from openpyxl.styles import Font

router = APIRouter(tags=["inventory"])


@router.get("/inventory/items")
def list_inventory(
    item_type: str = Query(None),
    kind: str = Query(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if kind == "chem":
        products = db.query(ChemProduct).order_by(ChemProduct.trade_name.asc()).all()
        payload = []
        for p in products:
            actives = (
                db.query(ProductActive, ActiveSubstance)
                .join(ActiveSubstance, ActiveSubstance.id == ProductActive.active_id)
                .filter(ProductActive.product_id == p.id)
                .all()
            )
            payload.append(
                {
                    "id": p.id,
                    "trade_name": p.trade_name,
                    "product_type": p.product_type,
                    "formulation": p.formulation,
                    "supplier": p.supplier,
                    "ean13": p.ean13,
                    "registration_no": p.registration_no,
                    "label_doc_id": p.label_doc_id,
                    "sds_doc_id": p.sds_doc_id,
                    "density_kg_per_l": p.density_kg_per_l,
                    "default_uom": p.default_uom,
                    "compat_notes": p.compat_notes,
                    "notes": p.notes,
                    "actives": [
                        {
                            "active_id": a.id,
                            "active_name": a.name,
                            "concentration": pa.concentration,
                            "unit": pa.unit,
                        }
                        for pa, a in actives
                    ],
                }
            )
        return payload

    query = db.query(Inventory)
    if item_type:
        query = query.filter(Inventory.item_type == item_type)
    return query.all()


@router.get("/inventory/actives")
def list_actives(db: Session = Depends(get_db), user=Depends(get_current_user)):
    return db.query(ActiveSubstance).order_by(ActiveSubstance.name.asc()).all()


@router.post("/inventory/actives")
def create_active(payload: ActiveSubstanceCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    canonical = chem_parse.normalize_active_name(payload.name)
    name_norm = chem_parse.normalize_text(canonical).replace(" ", "")
    existing = db.query(ActiveSubstance).filter(ActiveSubstance.name_norm == name_norm).first()
    if existing:
        existing.cas_no = payload.cas_no or existing.cas_no
        existing.synonyms = payload.synonyms or existing.synonyms
        existing.notes = payload.notes or existing.notes
        db.commit()
        db.refresh(existing)
        return existing
    active = ActiveSubstance(
        name=canonical,
        name_norm=name_norm,
        cas_no=payload.cas_no,
        synonyms=payload.synonyms,
        notes=payload.notes,
    )
    db.add(active)
    db.commit()
    db.refresh(active)
    return active


@router.patch("/inventory/actives/{active_id}")
def update_active(active_id: int, payload: ActiveSubstanceCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    active = db.query(ActiveSubstance).filter(ActiveSubstance.id == active_id).first()
    if not active:
        raise HTTPException(status_code=404, detail="Substanta activa nu exista")
    if payload.name:
        canonical = chem_parse.normalize_active_name(payload.name)
        active.name = canonical
        active.name_norm = chem_parse.normalize_text(canonical).replace(" ", "")
    active.cas_no = payload.cas_no
    active.synonyms = payload.synonyms
    active.notes = payload.notes
    db.commit()
    db.refresh(active)
    return active


@router.delete("/inventory/actives/{active_id}")
def delete_active(active_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    active = db.query(ActiveSubstance).filter(ActiveSubstance.id == active_id).first()
    if not active:
        raise HTTPException(status_code=404, detail="Substanta activa nu exista")
    usage = db.query(ProductActive).filter(ProductActive.active_id == active_id).first()
    if usage:
        raise HTTPException(status_code=400, detail="Substanta este folosita in produse")
    db.delete(active)
    db.commit()
    return {"ok": True}


@router.get("/inventory/products")
def list_products(
    product_type: str = Query(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    query = db.query(ChemProduct)
    if product_type:
        query = query.filter(ChemProduct.product_type == product_type)
    products = query.order_by(ChemProduct.trade_name.asc()).all()
    return [_product_to_dict(db, p) for p in products]


@router.get("/inventory/products/{product_id}")
def get_product(product_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    product = db.query(ChemProduct).filter(ChemProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produs inexistent")
    return _product_to_dict(db, product)


@router.post("/inventory/products")
def create_product(payload: ChemProductUpsert, db: Session = Depends(get_db), user=Depends(get_current_user)):
    return _upsert_chem_product(payload, db)


@router.patch("/inventory/products/{product_id}")
def update_product(product_id: int, payload: ChemProductUpsert, db: Session = Depends(get_db), user=Depends(get_current_user)):
    payload_data = payload.dict()
    payload_data["id"] = product_id
    return _upsert_chem_product(payload, db, product_id=product_id)


@router.delete("/inventory/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    product = db.query(ChemProduct).filter(ChemProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produs inexistent")
    has_lots = db.query(StockLot).filter(StockLot.product_id == product_id).first()
    if has_lots:
        raise HTTPException(status_code=400, detail="Produsul are loturi asociate")
    db.query(ProductActive).filter(ProductActive.product_id == product_id).delete()
    db.delete(product)
    db.commit()
    return {"ok": True}


@router.post("/inventory/items")
def create_inventory_item(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if isinstance(payload, dict) and ("trade_name" in payload or "actives" in payload):
        try:
            product_payload = ChemProductUpsert(**payload)
        except ValidationError as exc:
            raise HTTPException(status_code=422, detail=exc.errors())
        return _upsert_chem_product(product_payload, db)

    try:
        item_payload = InventoryCreate(**payload)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=exc.errors())
    item = Inventory(**item_payload.dict())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/inventory/items/{item_id}")
def update_inventory(item_id: int, payload: InventoryUpdate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    item = db.query(Inventory).filter(Inventory.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    data = payload.dict(exclude_unset=True)
    for key, value in data.items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@router.post("/inventory/movements")
def create_movement(payload: InventoryMovementCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    item = db.query(Inventory).filter(Inventory.id == payload.inventory_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    movement = InventoryMovement(**payload.dict())
    if movement.movement == "in":
        item.qty = (item.qty or 0) + movement.qty
    else:
        item.qty = (item.qty or 0) - movement.qty
    db.add(movement)
    db.commit()
    db.refresh(movement)
    return movement


@router.post("/inventory/ingest-label")
async def ingest_label(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    content = await file.read()
    ocr_endpoint = os.getenv("OCR_ENDPOINT", "")
    if not ocr_endpoint:
        raise HTTPException(status_code=500, detail="OCR service not configured")
    last_error = None
    resp = None
    for _ in range(5):
        try:
            resp = requests.post(
                f"{ocr_endpoint}/ocr",
                files={"file": (file.filename, content, file.content_type or "application/octet-stream")},
                timeout=90,
            )
            resp.raise_for_status()
            break
        except requests.RequestException as exc:
            last_error = exc
            time.sleep(2)
            resp = None
    if resp is None:
        raise HTTPException(status_code=503, detail=f"OCR indisponibil: {last_error}")
    data = resp.json()
    lines = [l.get("text", "") for l in data.get("lines", [])]
    parsed = chem_parse.parse_label_lines(lines)
    mapped_actives = chem_parse.map_actives_to_canonical(parsed.get("actives", []))

    for item in mapped_actives:
        name_norm = chem_parse.normalize_text(item.get("name", "")).replace(" ", "")
        active = db.query(ActiveSubstance).filter(ActiveSubstance.name_norm == name_norm).first()
        if not active:
            active = (
                db.query(ActiveSubstance)
                .filter(ActiveSubstance.synonyms.any(item.get("name")))
                .first()
            )
        if active:
            item["active_id"] = active.id
            item["name"] = active.name

    parsed["actives"] = mapped_actives

    product_match = None
    if parsed.get("trade_name"):
        trade_norm = chem_parse.normalize_text(parsed["trade_name"]).replace(" ", "")
        product = db.query(ChemProduct).filter(ChemProduct.trade_name_norm == trade_norm).first()
        if product:
            product_match = _product_to_dict(db, product)

    doc_key = storage.save_doc(content, file.filename, file.content_type or "application/octet-stream")
    doc = Doc(path=doc_key, type="label", ocr_json=str(data))
    db.add(doc)
    db.commit()

    return {
        "doc_id": doc.id,
        "product_suggestion": product_match,
        "actives_suggestion": parsed.get("actives", []),
        "ean13": parsed.get("ean13"),
        "raw_lines": parsed.get("raw_lines", []),
    }


@router.post("/inventory/lots")
def create_lot(payload: InventoryLotCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    product = db.query(ChemProduct).filter(ChemProduct.id == payload.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produsul nu exista")

    expires_at = payload.expiry_date or payload.expires_at
    if expires_at and expires_at < payload.received_date:
        raise HTTPException(status_code=400, detail="Data expirarii trebuie sa fie dupa data receptiei")

    uom = chem_units.normalize_uom(payload.uom)
    if uom not in chem_units.ALLOWED_UOM:
        raise HTTPException(status_code=400, detail="Unitatea trebuie sa fie l sau kg")

    if payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Cantitatea trebuie sa fie > 0")

    actives = db.query(ProductActive).filter(ProductActive.product_id == product.id).all()
    for pa in actives:
        unit = chem_units.normalize_conc_unit(pa.unit) or pa.unit
        if unit == "%w/v" and not product.density_kg_per_l:
            raise HTTPException(status_code=400, detail="Densitatea este obligatorie pentru produse %w/v")
        if chem_units.requires_density(unit, uom) and not product.density_kg_per_l:
            raise HTTPException(status_code=400, detail="Densitatea produsului este necesara pentru conversia unitatilor")

    location = _ensure_location(db, payload.location_id, payload.location_name)

    lot = StockLot(
        product_id=product.id,
        location_id=location.id,
        lot_code=payload.lot_code.strip(),
        received_date=payload.received_date,
        expires_at=expires_at,
        uom=uom,
        unit_price=payload.unit_price,
        notes=payload.notes,
    )
    db.add(lot)
    db.flush()

    txn = InventoryTxn(
        lot_id=lot.id,
        movement="in",
        qty=payload.qty,
        uom=uom,
        date=payload.received_date,
    )
    db.add(txn)
    db.commit()
    db.refresh(lot)

    expires_in = (expires_at - date.today()).days if expires_at else None
    return {
        "id": lot.id,
        "product_id": lot.product_id,
        "location_id": lot.location_id,
        "lot_code": lot.lot_code,
        "received_date": lot.received_date,
        "expires_at": lot.expires_at,
        "uom": lot.uom,
        "unit_price": lot.unit_price,
        "qty": payload.qty,
        "expires_in_days": expires_in,
        "expiring_soon": (expires_in is not None and expires_in <= 90),
    }


@router.get("/inventory/lots")
def list_lots(
    product_id: int = Query(None),
    location_id: int = Query(None),
    expiring_within_days: int = Query(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    params = {"product_id": product_id, "location_id": location_id}
    sql = """
        SELECT l.id AS lot_id,
               l.product_id,
               l.location_id,
               l.lot_code,
               l.received_date,
               l.expires_at,
               l.uom,
               l.unit_price,
               l.notes,
               COALESCE(SUM(CASE WHEN t.movement='in' THEN t.qty WHEN t.movement='out' THEN -t.qty ELSE t.qty END),0) AS qty
        FROM stock_lots l
        LEFT JOIN inventory_txns t ON t.lot_id = l.id
        WHERE (:product_id IS NULL OR l.product_id = :product_id)
          AND (:location_id IS NULL OR l.location_id = :location_id)
        GROUP BY l.id
        ORDER BY (l.expires_at IS NULL) ASC, l.expires_at ASC, l.received_date ASC;
    """
    rows = db.execute(text(sql), params).mappings().all()
    today = date.today()
    results = []
    for r in rows:
        expires_in = (r["expires_at"] - today).days if r["expires_at"] else None
        if expiring_within_days is not None and (expires_in is None or expires_in > expiring_within_days):
            continue
        results.append(
            {
                "id": r["lot_id"],
                "product_id": r["product_id"],
                "location_id": r["location_id"],
                "lot_code": r["lot_code"],
                "received_date": r["received_date"],
                "expires_at": r["expires_at"],
                "uom": r["uom"],
                "unit_price": r["unit_price"],
                "notes": r["notes"],
                "qty": float(r["qty"]),
                "expires_in_days": expires_in,
                "expiring_soon": (expires_in is not None and expires_in <= 90),
            }
        )
    return results


@router.post("/inventory/txns")
def create_inventory_txn(payload: InventoryTxnCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    movement = payload.movement.lower().strip()
    if movement not in {"in", "out", "adjust"}:
        raise HTTPException(status_code=400, detail="Movement invalid")

    uom = chem_units.normalize_uom(payload.uom)
    if uom not in chem_units.ALLOWED_UOM:
        raise HTTPException(status_code=400, detail="Unitatea trebuie sa fie l sau kg")

    if movement != "adjust" and payload.qty <= 0:
        raise HTTPException(status_code=400, detail="Cantitatea trebuie sa fie > 0")

    if payload.lot_id:
        lot = db.query(StockLot).filter(StockLot.id == payload.lot_id).first()
        if not lot:
            raise HTTPException(status_code=404, detail="Lotul nu exista")
        if lot.uom != uom:
            raise HTTPException(status_code=400, detail="Unitatea nu corespunde lotului")
        if movement == "out":
            current = _get_lot_balance(db, lot.id)
            if current < payload.qty:
                raise HTTPException(status_code=400, detail="Stoc insuficient in lot")
        txn = InventoryTxn(
            lot_id=lot.id,
            movement=movement,
            qty=payload.qty,
            uom=uom,
            date=payload.date,
            doc_id=payload.doc_id,
            reason=payload.reason,
            ref_type=payload.ref_type,
            ref_id=payload.ref_id if hasattr(payload, "ref_id") else None,
            created_by=payload.created_by,
            notes=payload.notes,
        )
        db.add(txn)
        db.commit()
        db.refresh(txn)
        return txn

    if movement != "out":
        raise HTTPException(status_code=400, detail="lot_id este obligatoriu pentru intrari/ajustari")
    if not payload.product_id:
        raise HTTPException(status_code=400, detail="product_id este obligatoriu pentru FIFO")

    allocations = _allocate_fifo(db, payload.product_id, payload.qty, uom)
    if not allocations:
        raise HTTPException(status_code=400, detail="Stoc insuficient pentru FIFO")

    created = []
    for alloc in allocations:
        txn = InventoryTxn(
            lot_id=alloc["lot_id"],
            movement="out",
            qty=alloc["qty"],
            uom=uom,
            date=payload.date,
            doc_id=payload.doc_id,
            notes=payload.notes,
        )
        db.add(txn)
        created.append(txn)
    db.commit()
    return {"items": [{"lot_id": t.lot_id, "qty": t.qty} for t in created]}


@router.get("/inventory/txns")
def list_inventory_txns(db: Session = Depends(get_db), user=Depends(get_current_user)):
    rows = db.query(InventoryTxn).order_by(InventoryTxn.created_at.desc()).all()
    return rows


@router.get("/inventory/active-stock")
def active_stock(active: str = Query(None), name: str = Query(None), db: Session = Depends(get_db), user=Depends(get_current_user)):
    query_name = name or active
    if not query_name:
        raise HTTPException(status_code=400, detail="name este obligatoriu")
    canonical = chem_parse.normalize_active_name(query_name)
    name_norm = chem_parse.normalize_text(canonical).replace(" ", "")
    active_row = db.query(ActiveSubstance).filter(ActiveSubstance.name_norm == name_norm).first()
    if not active_row:
        active_row = (
            db.query(ActiveSubstance)
            .filter(ActiveSubstance.synonyms.any(canonical))
            .first()
        )
    if not active_row:
        raise HTTPException(status_code=404, detail="Substanta activa nu exista")

    rows = db.execute(
        text(
            """
            SELECT active_id, active_name, product_id, trade_name, lot_id, lot_code, expires_at, uom, lot_qty, active_kg
            FROM vw_active_stock
            WHERE active_id = :active_id
            """
        ),
        {"active_id": active_row.id},
    ).mappings().all()

    total = 0.0
    breakdown = []
    for r in rows:
        qty = float(r["lot_qty"] or 0)
        if qty <= 0:
            continue
        active_kg = r["active_kg"]
        if active_kg is not None:
            total += float(active_kg)
        breakdown.append(
            {
                "product_id": r["product_id"],
                "trade_name": r["trade_name"],
                "lot_id": r["lot_id"],
                "lot_code": r["lot_code"],
                "qty": qty,
                "uom": r["uom"],
                "active_kg": float(active_kg) if active_kg is not None else None,
                "expires_at": r["expires_at"],
            }
        )

    return {"active_name": active_row.name, "total_kg": round(total, 3), "breakdown": breakdown}


@router.get("/inventory/stock-summary")
def stock_summary(db: Session = Depends(get_db), user=Depends(get_current_user)):
    sql = """
        SELECT l.product_id, l.location_id, l.uom,
               COALESCE(SUM(CASE WHEN t.movement='in' THEN t.qty WHEN t.movement='out' THEN -t.qty ELSE t.qty END),0) AS qty,
               COALESCE(SUM(CASE WHEN t.movement='in' THEN t.qty WHEN t.movement='out' THEN -t.qty ELSE t.qty END),0) * COALESCE(l.unit_price,0) AS value
        FROM stock_lots l
        LEFT JOIN inventory_txns t ON t.lot_id = l.id
        GROUP BY l.product_id, l.location_id, l.uom, l.unit_price
    """
    rows = db.execute(text(sql)).mappings().all()
    return [dict(r) for r in rows]


@router.get("/inventory/export.xlsx")
def export_excel(
    scope: str = Query("all"),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(name: str, headers: list[str], rows: list[list]):
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(row)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    if scope in {"products", "all"}:
        products = db.query(ChemProduct).order_by(ChemProduct.trade_name.asc()).all()
        rows = [
            [
                p.id,
                p.trade_name,
                p.formulation,
                p.supplier,
                p.ean13,
                p.registration_no,
                p.density_kg_per_l,
            ]
            for p in products
        ]
        add_sheet("Products", ["id", "trade_name", "formulation", "supplier", "ean13", "registration_no", "density"], rows)

    if scope in {"actives", "all"}:
        actives = db.query(ActiveSubstance).order_by(ActiveSubstance.name.asc()).all()
        rows = [
            [a.id, a.name, a.cas_no, ",".join(a.synonyms or []), a.notes]
            for a in actives
        ]
        add_sheet("ActiveSubstances", ["id", "name", "cas_no", "synonyms", "notes"], rows)

    if scope in {"product_actives", "all"}:
        rows = db.execute(
            text(
                """
                SELECT pa.id, pa.product_id, p.trade_name, pa.active_id, a.name, pa.concentration, pa.unit
                FROM product_actives pa
                JOIN chem_products p ON p.id = pa.product_id
                JOIN active_substances a ON a.id = pa.active_id
                """
            )
        ).fetchall()
        add_sheet("ProductActives", ["id", "product_id", "product", "active_id", "active", "concentration", "unit"], rows)

    if scope in {"lots", "all"}:
        rows = db.execute(
            text(
                """
                SELECT l.id, l.product_id, p.trade_name, l.lot_code, l.received_date, l.expires_at, l.uom, l.unit_price
                FROM stock_lots l
                JOIN chem_products p ON p.id = l.product_id
                ORDER BY (l.expires_at IS NULL) ASC, l.expires_at ASC
                """
            )
        ).fetchall()
        add_sheet("Lots", ["id", "product_id", "product", "lot_code", "received_date", "expires_at", "uom", "unit_price"], rows)

    if scope in {"txns", "all"}:
        rows = db.execute(
            text(
                """
                SELECT id, lot_id, movement, qty, uom, date, reason, ref_type, ref_id, created_at
                FROM inventory_txns
                ORDER BY created_at DESC
                """
            )
        ).fetchall()
        add_sheet("Transactions", ["id", "lot_id", "movement", "qty", "uom", "date", "reason", "ref_type", "ref_id", "created_at"], rows)

    if scope in {"applications", "all"}:
        rows = db.execute(
            text(
                """
                SELECT a.id, a.parcel_id, a.date, a.area_ha, a.total_cost
                FROM applications a
                ORDER BY a.date DESC
                """
            )
        ).fetchall()
        add_sheet("Applications", ["id", "parcel_id", "date", "area_ha", "total_cost"], rows)

    if scope in {"active_summary", "all"}:
        rows = db.execute(
            text(
                """
                SELECT active_name, SUM(active_kg) AS total_kg
                FROM vw_active_stock
                GROUP BY active_name
                ORDER BY active_name ASC
                """
            )
        ).fetchall()
        add_sheet("ActiveStockSummary", ["active_name", "total_kg"], rows)

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_export.xlsx"},
    )


@router.post("/ocr/label")
async def ocr_label(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    content = await file.read()
    ocr_endpoint = os.getenv("OCR_ENDPOINT", "")
    if not ocr_endpoint:
        raise HTTPException(status_code=500, detail="OCR service not configured")
    resp = requests.post(
        f"{ocr_endpoint}/ocr",
        files={"file": (file.filename, content, file.content_type or "application/octet-stream")},
        timeout=60,
    )
    resp.raise_for_status()
    data = resp.json()
    lines = [l.get("text", "") for l in data.get("lines", [])]
    parsed = chem_parse.parse_label_lines(lines)

    doc_key = storage.save_doc(content, file.filename, file.content_type or "application/octet-stream")
    doc = Doc(path=doc_key, type="label", ocr_json=str(data))
    db.add(doc)
    db.commit()

    return {"doc_id": doc.id, "parsed": parsed}


@router.post("/mix/check")
def check_mix(payload: MixCheckRequest, db: Session = Depends(get_db), user=Depends(get_current_user)):
    a = payload.a_subst
    b = payload.b_subst
    rule = db.query(ChemMixRule).filter(ChemMixRule.a_subst == a, ChemMixRule.b_subst == b).first()
    if not rule:
        rule = db.query(ChemMixRule).filter(ChemMixRule.a_subst == b, ChemMixRule.b_subst == a).first()
    if not rule:
        return {"status": "unknown"}
    if rule.allowed:
        return {"status": "allowed", "notes": rule.notes}
    return {"status": "forbidden", "notes": rule.notes}


@router.post("/mix/check-items")
def check_mix_items(
    items: list[dict],
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    # items: [{product_id, dose_per_ha, uom}]
    product_ids = [int(i.get("product_id")) for i in items if i.get("product_id")]
    if not product_ids:
        raise HTTPException(status_code=400, detail="Nu exista produse pentru verificare")

    actives = (
        db.query(ProductActive, ActiveSubstance)
        .join(ActiveSubstance, ActiveSubstance.id == ProductActive.active_id)
        .filter(ProductActive.product_id.in_(product_ids))
        .all()
    )
    active_names = sorted({a.name for _, a in actives})
    pairs = []
    worst = "allowed"

    def severity(val: str) -> int:
        order = {"forbidden": 3, "caution": 2, "unknown": 1, "allowed": 0}
        return order.get(val, 1)

    for i in range(len(active_names)):
        for j in range(i + 1, len(active_names)):
            a = active_names[i]
            b = active_names[j]
            rule = db.query(ChemMixRule).filter(ChemMixRule.a_subst == a, ChemMixRule.b_subst == b).first()
            if not rule:
                rule = db.query(ChemMixRule).filter(ChemMixRule.a_subst == b, ChemMixRule.b_subst == a).first()
            if not rule:
                relation = "unknown"
                note = None
            else:
                relation = "allowed" if rule.allowed else "forbidden"
                note = rule.notes
            pairs.append({"a": a, "b": b, "relation": relation, "note": note})
            if severity(relation) > severity(worst):
                worst = relation

    return {"summary": worst, "pairs": pairs}


def _upsert_chem_product(payload: ChemProductUpsert, db: Session, product_id: int | None = None):
    trade_norm = chem_parse.normalize_text(payload.trade_name).replace(" ", "")
    product = None
    if product_id:
        product = db.query(ChemProduct).filter(ChemProduct.id == product_id).first()
    if not product:
        product = db.query(ChemProduct).filter(ChemProduct.trade_name_norm == trade_norm).first()
    if not product:
        product = ChemProduct(trade_name=payload.trade_name, trade_name_norm=trade_norm)
        db.add(product)
        db.flush()

    product.trade_name = payload.trade_name
    product.trade_name_norm = trade_norm
    product.product_type = (payload.product_type or product.product_type or "herbicide").lower()
    product.formulation = payload.formulation
    product.supplier = payload.supplier
    product.ean13 = payload.ean13
    product.registration_no = payload.registration_no
    product.density_kg_per_l = payload.density_kg_per_l
    product.default_uom = chem_units.normalize_uom(payload.default_uom) if payload.default_uom else None
    product.label_doc_id = payload.label_doc_id
    product.sds_doc_id = payload.sds_doc_id
    product.compat_notes = payload.compat_notes
    product.notes = payload.notes

    if payload.actives:
        for active_input in payload.actives:
            unit = chem_units.normalize_conc_unit(active_input.unit)
            if unit not in chem_units.ALLOWED_CONC_UNITS:
                raise HTTPException(status_code=400, detail="Unitatea concentratiei nu este acceptata")
            err = chem_units.validate_concentration(active_input.concentration, unit)
            if err:
                raise HTTPException(status_code=400, detail=err)
            if unit == "%w/v" and not product.density_kg_per_l:
                raise HTTPException(status_code=400, detail="Densitatea este obligatorie pentru produse %w/v")

            active = _resolve_active(db, active_input.active_id, active_input.active_name)
            existing = (
                db.query(ProductActive)
                .filter(ProductActive.product_id == product.id, ProductActive.active_id == active.id)
                .first()
            )
            if existing:
                existing.concentration = active_input.concentration
                existing.unit = unit
            else:
                db.add(
                    ProductActive(
                        product_id=product.id,
                        active_id=active.id,
                        concentration=active_input.concentration,
                        unit=unit,
                    )
                )

    db.commit()
    db.refresh(product)

    actives = (
        db.query(ProductActive, ActiveSubstance)
        .join(ActiveSubstance, ActiveSubstance.id == ProductActive.active_id)
        .filter(ProductActive.product_id == product.id)
        .all()
    )

    return {
        "id": product.id,
        "trade_name": product.trade_name,
        "product_type": product.product_type,
        "formulation": product.formulation,
        "density_kg_per_l": product.density_kg_per_l,
        "default_uom": product.default_uom,
        "notes": product.notes,
        "actives": [
            {
                "active_id": a.id,
                "active_name": a.name,
                "concentration": pa.concentration,
                "unit": pa.unit,
            }
            for pa, a in actives
        ],
    }


def _resolve_active(db: Session, active_id: int | None, active_name: str | None) -> ActiveSubstance:
    if active_id:
        active = db.query(ActiveSubstance).filter(ActiveSubstance.id == active_id).first()
        if not active:
            raise HTTPException(status_code=404, detail="Substanta activa nu exista")
        return active
    if not active_name:
        raise HTTPException(status_code=400, detail="active_name este obligatoriu")
    canonical = chem_parse.normalize_active_name(active_name)
    name_norm = chem_parse.normalize_text(canonical).replace(" ", "")
    active = db.query(ActiveSubstance).filter(ActiveSubstance.name_norm == name_norm).first()
    if active:
        return active
    active = db.query(ActiveSubstance).filter(ActiveSubstance.synonyms.any(canonical)).first()
    if active:
        return active
    active = ActiveSubstance(name=canonical, name_norm=name_norm)
    db.add(active)
    db.flush()
    return active


def _ensure_location(db: Session, location_id: int | None, location_name: str | None) -> InventoryLocation:
    if location_id:
        location = db.query(InventoryLocation).filter(InventoryLocation.id == location_id).first()
        if not location:
            raise HTTPException(status_code=404, detail="Locatia nu exista")
        return location
    name = (location_name or "Depozit principal").strip()
    location = db.query(InventoryLocation).filter(InventoryLocation.name == name).first()
    if location:
        return location
    location = InventoryLocation(name=name)
    db.add(location)
    db.flush()
    return location


def _product_to_dict(db: Session, product: ChemProduct):
    actives = (
        db.query(ProductActive, ActiveSubstance)
        .join(ActiveSubstance, ActiveSubstance.id == ProductActive.active_id)
        .filter(ProductActive.product_id == product.id)
        .all()
    )
    return {
        "id": product.id,
        "trade_name": product.trade_name,
        "product_type": product.product_type,
        "formulation": product.formulation,
        "supplier": product.supplier,
        "ean13": product.ean13,
        "registration_no": product.registration_no,
        "label_doc_id": product.label_doc_id,
        "sds_doc_id": product.sds_doc_id,
        "density_kg_per_l": product.density_kg_per_l,
        "default_uom": product.default_uom,
        "compat_notes": product.compat_notes,
        "notes": product.notes,
        "actives": [
            {
                "active_id": a.id,
                "active_name": a.name,
                "concentration": pa.concentration,
                "unit": pa.unit,
            }
            for pa, a in actives
        ],
    }


def _get_lot_balance(db: Session, lot_id: int) -> float:
    sql = """
        SELECT COALESCE(SUM(CASE WHEN movement='in' THEN qty WHEN movement='out' THEN -qty ELSE qty END),0) AS qty
        FROM inventory_txns
        WHERE lot_id = :lot_id
    """
    row = db.execute(text(sql), {"lot_id": lot_id}).mappings().first()
    return float(row["qty"] or 0)


def _allocate_fifo(db: Session, product_id: int, qty: float, uom: str):
    sql = """
        SELECT l.id AS lot_id,
               l.uom,
               l.expires_at,
               l.received_date,
               COALESCE(SUM(CASE WHEN t.movement='in' THEN t.qty WHEN t.movement='out' THEN -t.qty ELSE t.qty END),0) AS qty
        FROM stock_lots l
        LEFT JOIN inventory_txns t ON t.lot_id = l.id
        WHERE l.product_id = :product_id
        GROUP BY l.id
        ORDER BY (l.expires_at IS NULL) ASC, l.expires_at ASC, l.received_date ASC, l.id ASC
    """
    rows = db.execute(text(sql), {"product_id": product_id}).mappings().all()

    remaining = qty
    allocations = []
    for r in rows:
        if remaining <= 0:
            break
        if r["uom"] != uom:
            continue
        available = float(r["qty"] or 0)
        if available <= 0:
            continue
        take = available if available < remaining else remaining
        allocations.append({"lot_id": r["lot_id"], "qty": take})
        remaining -= take

    if remaining > 0:
        return []
    return allocations
